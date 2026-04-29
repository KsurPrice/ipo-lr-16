from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import EmailMessage
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import Order, OrderItem
from cart.models import Cart

@login_required
def checkout(request):
    """Страница оформления заказа"""
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all().select_related('product')
    
    if not cart_items:
        messages.error(request, 'Ваша корзина пуста!')
        return redirect('products:product_list')
    
    if request.method == 'POST':
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        
        if not all([address, phone, email]):
            messages.error(request, 'Пожалуйста, заполните все поля!')
            return render(request, 'orders/checkout.html', {
                'cart_items': cart_items,
                'total_price': cart.get_total_price()
            })
        
        # Создаем заказ
        order = Order.objects.create(
            user=request.user,
            address=address,
            phone=phone,
            email=email,
            total_price=cart.get_total_price()
        )
        
        # Копируем товары из корзины в заказ
        for item in cart_items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )
        
        # Генерируем Excel чек
        excel_file = generate_receipt(order)
        
        # Отправляем чек по email
        send_receipt_email(order, excel_file)
        
        # Очищаем корзину
        cart.items.all().delete()
        
        messages.success(request, f'Заказ #{order.id} успешно оформлен! Чек отправлен на {email}')
        return redirect('products:product_list')
    
    return render(request, 'orders/checkout.html', {
        'cart_items': cart_items,
        'total_price': cart.get_total_price(),
    })

def generate_receipt(order):
    """Генерация чека в формате Excel"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Чек_заказа_{order.id}"
    
    # Стили
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f'ЧЕК ЗАКАЗА №{order.id}'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Информация о заказе
    ws['A3'] = 'Дата заказа:'
    ws['B3'] = order.created_at.strftime('%d.%m.%Y %H:%M')
    ws['A4'] = 'Покупатель:'
    ws['B4'] = order.user.username
    ws['A5'] = 'Адрес доставки:'
    ws['B5'] = order.address
    ws['A6'] = 'Телефон:'
    ws['B6'] = order.phone
    ws['A7'] = 'Email:'
    ws['B7'] = order.email
    
    # Заголовки таблицы
    headers = ['№', 'Товар', 'Цена', 'Количество', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Товары
    row = 10
    for idx, item in enumerate(order.items.all(), 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=item.product.name).border = thin_border
        ws.cell(row=row, column=3, value=float(item.price)).border = thin_border
        ws.cell(row=row, column=4, value=item.quantity).border = thin_border
        ws.cell(row=row, column=5, value=float(item.get_total())).border = thin_border
        row += 1
    
    # Итого
    ws.cell(row=row, column=4, value='ИТОГО:').font = header_font
    ws.cell(row=row, column=5, value=float(order.total_price)).font = header_font
    ws.merge_cells(f'A{row}:D{row}')
    
    # Автоподбор ширины колонок (ИСПРАВЛЕННАЯ ВЕРСИЯ)
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            try:
                # Пропускаем объединенные ячейки
                if isinstance(cell, (str, int, float)):
                    cell_value = str(cell)
                else:
                    cell_value = str(cell.value) if cell.value else ''
                
                if len(cell_value) > column_widths.get(cell.column, 0):
                    column_widths[cell.column] = min(len(cell_value) + 2, 50)
            except:
                pass
    
    for col, width in column_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width
    
    wb.save(output)
    output.seek(0)
    return output

def send_receipt_email(order, excel_file):
    """Отправка чека по email"""
    subject = f'Чек заказа №{order.id} в BikeShop'
    message = f'''
    Здравствуйте, {order.user.username}!
    
    Спасибо за покупку в нашем магазине велоаксессуаров!
    
    Ваш заказ №{order.id} успешно оформлен.
    Сумма заказа: {order.total_price} руб.
    Дата заказа: {order.created_at.strftime('%d.%m.%Y %H:%M')}
    
    Чек прикреплен к этому письму.
    
    С уважением,
    Команда BikeShop
    '''
    
    email_message = EmailMessage(
        subject,
        message,
        'noreply@bikeshop.com',
        [order.email],
    )
    
    email_message.attach(
        f'receipt_order_{order.id}.xlsx',
        excel_file.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    email_message.send()